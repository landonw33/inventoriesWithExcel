import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
products_under_ten_inv = {}

# print(product_list.max_row)
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row, 1).value
    inv_price = product_list.cell(product_row, 5)
    # print(supplier_name)

    # calculation number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier[supplier_name]
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        # print("adding a new supplier")
        products_per_supplier[supplier_name] = 1

    # find total inv per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + (inventory * price)
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # when the inv is less than 10
    if inventory < 10:
        products_under_ten_inv[int(product_number)] = int(inventory)

    # add value for total inv price
    inv_price.value = inventory*price

# shows how much inv each supplier has
print(products_per_supplier)
# shows how much each suppliers' inv is worth
print(total_value_per_supplier)
# shows products with inv less than 10
print(products_under_ten_inv)
# saving to our excel sheet
inv_file.save("inventory_with_total_value.xlsx")
